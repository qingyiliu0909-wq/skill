import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
from openpyxl import load_workbook

HEADER_ROWS = 4
DATA_START_ROW = 5
EMPTY_VALUES = {"", "none", "null", "nil", "nan"}

def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() in EMPTY_VALUES:
        return ""
    return text

def is_empty(value):
    return normalize_cell(value) == ""

def value_at(row, index):
    if index >= len(row):
        return None
    return row[index]

def normalize_field_name(field_name_raw, desc):
    if not is_empty(field_name_raw):
        return normalize_cell(field_name_raw)
    if not is_empty(desc):
        return normalize_cell(desc)
    return ""

def derive_row_key(values, key_field):
    if key_field and not is_empty(values.get(key_field)):
        return normalize_cell(values[key_field])
    for value in values.values():
        if not is_empty(value):
            return normalize_cell(value)
    return "<empty>"

def load_sheet_table(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    print(f"Total rows: {len(rows)}")
    if len(rows) < HEADER_ROWS:
        return {}
    
    header_rows = rows[:HEADER_ROWS]
    columns = {}
    max_cols = max(len(row) for row in header_rows)
    print(f"Max cols: {max_cols}")
    
    for col_index in range(max_cols):
        desc = value_at(header_rows[0], col_index)
        field_name_raw = value_at(header_rows[2], col_index)
        if all(is_empty(v) for v in [desc, field_name_raw]):
            continue
        field_name = normalize_field_name(field_name_raw, desc)
        if field_name:
            columns[field_name] = col_index
            print(f"Column {col_index}: {field_name}")
    
    data_rows = {}
    key_field = next(iter(columns), None)
    print(f"Key field: {key_field}")
    
    for row_number, raw_row in enumerate(rows[HEADER_ROWS:HEADER_ROWS+10], start=DATA_START_ROW):
        values = {}
        for field_name, col_index in columns.items():
            values[field_name] = value_at(raw_row, col_index)
        
        if not any(not is_empty(v) for v in values.values()):
            continue
        
        row_key = derive_row_key(values, key_field)
        print(f"Row {row_number}: key={row_key}")
        
        if not row_key:
            continue
        
        data_rows[row_key] = {"row_number": row_number, "key": row_key}
    
    print(f"Total data rows parsed: {len(data_rows)}")
    return data_rows

# Test with Quest.xlsx
file_path = Path(r"D:\OBT1.4Geili\EM\ExportDatas\datas\Quest.xlsx")
print(f"Loading file: {file_path}")
book = load_workbook(file_path, read_only=True, data_only=True)
print(f"Sheets: {book.sheetnames}")

for sheet_name in book.sheetnames:
    print(f"\n=== Sheet: {sheet_name} ===")
    load_sheet_table(book[sheet_name])

book.close()