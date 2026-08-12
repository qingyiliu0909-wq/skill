import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
from openpyxl import load_workbook

HEADER_ROWS = 4
DATA_START_ROW = 5
EMPTY_VALUES = {"", "none", "null", "nil", "nan"}

def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() in EMPTY_VALUES:
        return ""
    return text

def is_empty(value):
    return normalize_cell(value) == ""

def value_at(row, index):
    if index >= len(row):
        return None
    return row[index]

def normalize_field_name(field_name_raw, desc):
    if not is_empty(field_name_raw):
        return normalize_cell(field_name_raw)
    if not is_empty(desc):
        return normalize_cell(desc)
    return ""

def derive_row_key(values, key_field):
    if key_field and not is_empty(values.get(key_field)):
        return normalize_cell(values[key_field])
    for value in values.values():
        if not is_empty(value):
            return normalize_cell(value)
    return "<empty>"

def find_key_field(columns):
    for field_name in columns:
        if 'Id' in field_name or 'ID' in field_name:
            return field_name
    return next(iter(columns), None)

def load_sheet_table(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < HEADER_ROWS:
        return {}
    
    header_rows = rows[:HEADER_ROWS]
    columns = {}
    max_cols = max(len(row) for row in header_rows)
    
    for col_index in range(max_cols):
        desc = value_at(header_rows[0], col_index)
        field_name_raw = value_at(header_rows[2], col_index)
        if all(is_empty(v) for v in [desc, field_name_raw]):
            continue
        field_name = normalize_field_name(field_name_raw, desc)
        if field_name:
            columns[field_name] = col_index
    
    data_rows = {}
    key_field = find_key_field(columns)
    print(f"Key field: {key_field}")
    
    for row_number, raw_row in enumerate(rows[HEADER_ROWS:], start=DATA_START_ROW):
        values = {}
        for field_name, col_index in columns.items():
            values[field_name] = value_at(raw_row, col_index)
        
        if not any(not is_empty(v) for v in values.values()):
            continue
        
        row_key = derive_row_key(values, key_field)
        data_rows[row_key] = {"row_number": row_number, "key": row_key}
    
    return data_rows

# Main
a_file = Path(r"C:\Pan01\demo\EM_Build\ExportDatas\datas\Quest.xlsx")
b_file = Path(r"D:\OBT1.4Geili\EM\ExportDatas\datas\Quest.xlsx")

a_book = load_workbook(a_file, read_only=True, data_only=True)
b_book = load_workbook(b_file, read_only=True, data_only=True)

a_rows = load_sheet_table(a_book['任务链表|QuestChain'])
b_rows = load_sheet_table(b_book['任务链表|QuestChain'])

print(f"\nA rows count: {len(a_rows)}")
print(f"B rows count: {len(b_rows)}")
print(f"\nA has 100405: {'100405' in a_rows}")
print(f"B has 100405: {'100405' in b_rows}")

# Find missing keys
missing_in_a = [k for k in b_rows if k not in a_rows]
print(f"\nMissing in A (B has but A doesn't): {len(missing_in_a)}")
if missing_in_a:
    print("First 20 missing keys:")
    for k in missing_in_a[:20]:
        print(f"  {k}")

a_book.close()
b_book.close()