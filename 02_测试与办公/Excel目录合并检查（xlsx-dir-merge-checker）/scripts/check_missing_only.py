from __future__ import annotations
import sys
import warnings
from dataclasses import dataclass
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
from openpyxl import load_workbook

HEADER_ROWS = 4
DATA_START_ROW = 5
EMPTY_VALUES = {"", "none", "null", "nil", "nan"}


@dataclass(slots=True)
class DiffItem:
    table: str
    sheet: str
    diff_type: str
    header: str
    detail: str


def normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() in EMPTY_VALUES:
        return ""
    return text


def is_empty(value) -> bool:
    return normalize_cell(value) == ""


def value_at(row, index):
    if index >= len(row):
        return None
    return row[index]


def normalize_field_name(field_name_raw, desc):
    if not is_empty(field_name_raw):
        return normalize_cell(field_name_raw)
    if not is_empty(desc):
        return normalize_cell(desc)
    return ""


def derive_row_key(values, key_field):
    if key_field and not is_empty(values.get(key_field)):
        return normalize_cell(values[key_field])
    for value in values.values():
        if not is_empty(value):
            return normalize_cell(value)
    return "<empty>"


def find_key_field(columns):
    """查找主键列，优先选择包含 'Id' 或 'ID' 的列"""
    for field_name in columns:
        if 'Id' in field_name or 'ID' in field_name:
            return field_name
    # 如果没有找到包含 Id 的列，返回第一列
    return next(iter(columns), None)


def load_sheet_table(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < HEADER_ROWS:
        return {}
    
    header_rows = rows[:HEADER_ROWS]
    columns = {}
    max_cols = max(len(row) for row in header_rows)
    
    for col_index in range(max_cols):
        desc = value_at(header_rows[0], col_index)
        field_name_raw = value_at(header_rows[2], col_index)
        if all(is_empty(v) for v in [desc, field_name_raw]):
            continue
        field_name = normalize_field_name(field_name_raw, desc)
        if field_name:
            columns[field_name] = col_index
    
    data_rows = {}
    key_field = find_key_field(columns)
    
    # 查找"不导出"列（可能的列名：Skip, skip, 不导出, 是否导出等）
    skip_column = None
    for field_name in columns:
        if 'Skip' in field_name or 'skip' in field_name or '不导出' in field_name or '是否导出' in field_name:
            skip_column = field_name
            break
    
    for row_number, raw_row in enumerate(rows[HEADER_ROWS:], start=DATA_START_ROW):
        values = {}
        for field_name, col_index in columns.items():
            values[field_name] = value_at(raw_row, col_index)
        
        if not any(not is_empty(v) for v in values.values()):
            continue
        
        # 如果有"不导出"列且该行标记为不导出，则跳过
        if skip_column and skip_column in values:
            skip_value = normalize_cell(values[skip_column])
            if skip_value and skip_value != '0' and skip_value.lower() != 'false':
                continue
        
        row_key = derive_row_key(values, key_field)
        if not row_key:
            continue
        
        data_rows[row_key] = {"row_number": row_number, "key": row_key}
    
    return data_rows


def compare_workbook_pair(a_file, b_file, table_name):
    diffs = []
    a_book = load_workbook(a_file, read_only=True, data_only=True)
    b_book = load_workbook(b_file, read_only=True, data_only=True)
    
    try:
        b_sheets = set(b_book.sheetnames)
        a_sheets = set(a_book.sheetnames)
        
        for sheet_name in b_sheets:
            if sheet_name not in a_sheets:
                diffs.append(DiffItem(
                    table=table_name,
                    sheet=sheet_name,
                    diff_type="missing_sheet",
                    header="*",
                    detail="目录 A 缺少此 Sheet"
                ))
                continue
            
            a_rows = load_sheet_table(a_book[sheet_name])
            b_rows = load_sheet_table(b_book[sheet_name])
            
            for row_key in b_rows:
                if row_key not in a_rows:
                    diffs.append(DiffItem(
                        table=table_name,
                        sheet=sheet_name,
                        diff_type="missing_row",
                        header=row_key,
                        detail=f"目录 A 缺少数据行，主键='{row_key}'，行号={b_rows[row_key]['row_number']}"
                    ))
    finally:
        a_book.close()
        b_book.close()
    
    return diffs


def compare_directories(dir_a, dir_b):
    diffs = []
    b_files = sorted(path for path in dir_b.glob("**/*.xlsx") if path.is_file() and not path.name.startswith("~$"))
    
    for b_file in b_files:
        rel_path = b_file.relative_to(dir_b)
        a_file = dir_a / rel_path
        table_name = rel_path.as_posix()
        
        if not a_file.exists():
            diffs.append(DiffItem(
                table=table_name,
                sheet="*",
                diff_type="missing_table",
                header="*",
                detail="目录 A 缺少此表"
            ))
            continue
        
        diffs.extend(compare_workbook_pair(a_file, b_file, table_name))
    
    return diffs


def format_report(diffs, dir_a, dir_b):
    lines = [
        "=" * 72,
        "缺失行检查报告",
        f"目录 A（目标）: {dir_a}",
        f"目录 B（来源）: {dir_b}",
        "=" * 72,
    ]
    
    if not diffs:
        lines.append("未发现目录 B 中有、但目录 A 缺失的内容。")
        lines.append("=" * 72)
        return "\n".join(lines)
    
    grouped = {}
    for item in diffs:
        grouped.setdefault(item.table, []).append(item)
    
    for table_name, items in grouped.items():
        lines.append(f"\n[表] {table_name}")
        for item in items:
            lines.append(f"  - 类型: {item.diff_type}")
            lines.append(f"    Sheet: {item.sheet}")
            lines.append(f"    主键: {item.header}")
            lines.append(f"    详情: {item.detail}")
    
    lines.append("\n" + "=" * 72)
    lines.append(f"共发现 {len(diffs)} 处缺失内容。")
    lines.append("=" * 72)
    return "\n".join(lines)


def main():
    if len(sys.argv) != 3:
        print("用法: python check_missing_only.py <目录A> <目录B>", file=sys.stderr)
        return 1
    
    dir_a = Path(sys.argv[1])
    dir_b = Path(sys.argv[2])
    
    diffs = compare_directories(dir_a, dir_b)
    report = format_report(diffs, dir_a, dir_b)
    
    print(report)
    
    # 同时写入文件
    with open("missing_only_result.txt", "w", encoding="utf-8") as f:
        f.write(report)
    
    return 1 if diffs else 0


if __name__ == "__main__":
    raise SystemExit(main())