from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import tempfile
import warnings
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

from openpyxl import load_workbook


HEADER_ROWS = 4
DATA_START_ROW = 5
EMPTY_VALUES = {"", "none", "null", "nil", "nan"}
SHEET_NAME_PATTERN = re.compile(r"【.*?[|｜](.+?)】")


@dataclass(slots=True)
class DiffItem:
    table: str
    sheet: str
    diff_type: str
    header: str
    detail: str
    svn_revision: str | None = None
    svn_author: str | None = None
    svn_message: str | None = None


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    dir_a = Path(args.dir_a).resolve()
    dir_b = Path(args.dir_b).resolve()
    if not dir_a.is_dir():
        print(f"目录 A 不存在：{dir_a}", file=sys.stderr)
        return 2
    if not dir_b.is_dir():
        print(f"目录 B 不存在：{dir_b}", file=sys.stderr)
        return 2

    diffs = compare_directories(
        dir_a=dir_a,
        dir_b=dir_b,
        glob_pattern=args.glob,
        svn_lookup=not args.no_svn,
        svn_log_limit=args.svn_log_limit,
    )
    if args.json:
        print(json.dumps([asdict(item) for item in diffs], ensure_ascii=False, indent=2), file=sys.stderr)
    else:
        print(format_report(diffs, dir_a, dir_b), file=sys.stderr)
    return 1 if diffs else 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compare directory B against directory A and report unmerged Excel table content."
    )
    parser.add_argument("dir_a", help="Target directory (merged destination, 目录 A)")
    parser.add_argument("dir_b", help="Source directory (should be merged into A, 目录 B)")
    parser.add_argument(
        "--glob",
        default="**/*.xlsx",
        help="Glob pattern relative to each root directory. Default: **/*.xlsx",
    )
    parser.add_argument(
        "--no-svn",
        action="store_true",
        help="Skip SVN revision lookup for each diff item.",
    )
    parser.add_argument(
        "--svn-log-limit",
        type=int,
        default=80,
        help="How many recent SVN revisions to inspect per file when attributing diffs.",
    )
    parser.add_argument("--json", action="store_true", help="Output machine-readable JSON.")
    return parser


def compare_directories(
    *,
    dir_a: Path,
    dir_b: Path,
    glob_pattern: str,
    svn_lookup: bool,
    svn_log_limit: int,
) -> list[DiffItem]:
    diffs: list[DiffItem] = []
    b_files = sorted(path for path in dir_b.glob(glob_pattern) if path.is_file() and not path.name.startswith("~$"))
    for b_file in b_files:
        rel_path = b_file.relative_to(dir_b)
        a_file = dir_a / rel_path
        table_name = rel_path.as_posix()
        if not a_file.exists():
            diffs.append(
                DiffItem(
                    table=table_name,
                    sheet="*",
                    diff_type="missing_table",
                    header="*",
                    detail="目录 B 存在该表，目录 A 不存在。",
                )
            )
            if svn_lookup:
                attach_svn_info(diffs[-1], b_file, svn_log_limit)
            continue
        diffs.extend(
            compare_workbook_pair(
                a_file=a_file,
                b_file=b_file,
                table_name=table_name,
                svn_lookup=svn_lookup,
                svn_log_limit=svn_log_limit,
            )
        )
    return diffs


def compare_workbook_pair(
    *,
    a_file: Path,
    b_file: Path,
    table_name: str,
    svn_lookup: bool,
    svn_log_limit: int,
) -> list[DiffItem]:
    diffs: list[DiffItem] = []
    a_book = load_workbook(a_file, read_only=True, data_only=True)
    b_book = load_workbook(b_file, read_only=True, data_only=True)
    try:
        b_sheet_map = {normalize_sheet_key(name): name for name in b_book.sheetnames}
        a_sheet_map = {normalize_sheet_key(name): name for name in a_book.sheetnames}

        for sheet_key, b_sheet_name in b_sheet_map.items():
            if sheet_key not in a_sheet_map:
                item = DiffItem(
                    table=table_name,
                    sheet=b_sheet_name,
                    diff_type="missing_sheet",
                    header="*",
                    detail="目录 B 存在该 Sheet，目录 A 不存在。",
                )
                diffs.append(item)
                if svn_lookup:
                    attach_svn_info(item, b_file, svn_log_limit)
                continue

            a_sheet_name = a_sheet_map[sheet_key]
            a_table = load_sheet_table(a_book[a_sheet_name])
            b_table = load_sheet_table(b_book[b_sheet_name])
            diffs.extend(
                compare_sheet_tables(
                    table_name=table_name,
                    sheet_name=b_sheet_name,
                    a_table=a_table,
                    b_table=b_table,
                    b_file=b_file,
                    svn_lookup=svn_lookup,
                    svn_log_limit=svn_log_limit,
                )
            )
    finally:
        a_book.close()
        b_book.close()
    return diffs


def compare_sheet_tables(
    *,
    table_name: str,
    sheet_name: str,
    a_table: SheetTable,
    b_table: SheetTable,
    b_file: Path,
    svn_lookup: bool,
    svn_log_limit: int,
) -> list[DiffItem]:
    diffs: list[DiffItem] = []

    for row_key, b_row in b_table.rows.items():
        if row_key not in a_table.rows:
            item = DiffItem(
                table=table_name,
                sheet=sheet_name,
                diff_type="missing_row",
                header=row_key,
                detail=f"目录 A 缺少数据行，主键={row_key!r}。",
            )
            diffs.append(item)
            if svn_lookup:
                attach_row_revision(item, b_file, sheet_name, row_key, svn_log_limit)
            continue

        a_row = a_table.rows[row_key]
        shared_fields = [name for name in b_table.columns if name in a_table.columns]
        for field_name in shared_fields:
            a_value = a_row.values.get(field_name)
            b_value = b_row.values.get(field_name)
            if normalize_cell(a_value) == normalize_cell(b_value):
                continue
            item = DiffItem(
                table=table_name,
                sheet=sheet_name,
                diff_type="cell_mismatch",
                header=f"{row_key} / {field_name}",
                detail=f"A={format_value(a_value)!r}，B={format_value(b_value)!r}",
            )
            diffs.append(item)
            if svn_lookup:
                attach_cell_revision(
                    item,
                    b_file,
                    sheet_name,
                    row_key,
                    field_name,
                    b_value,
                    svn_log_limit,
                )
    return diffs


@dataclass(slots=True)
class ColumnHeader:
    index: int
    desc: Any
    type_name: Any
    field_name: str
    export_target: Any


@dataclass(slots=True)
class DataRow:
    row_number: int
    key: str
    values: dict[str, Any]


@dataclass(slots=True)
class SheetTable:
    columns: dict[str, ColumnHeader]
    rows: dict[str, DataRow]


def load_sheet_table(worksheet) -> SheetTable:
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < HEADER_ROWS:
        return SheetTable(columns={}, rows={})

    header_rows = rows[:HEADER_ROWS]
    columns: dict[str, ColumnHeader] = {}
    max_cols = max(len(row) for row in header_rows)
    for col_index in range(max_cols):
        desc = value_at(header_rows[0], col_index)
        type_name = value_at(header_rows[1], col_index)
        field_name_raw = value_at(header_rows[2], col_index)
        export_target = value_at(header_rows[3], col_index)
        if all(is_empty(value) for value in (desc, type_name, field_name_raw, export_target)):
            continue
        field_name = normalize_field_name(field_name_raw, desc)
        if not field_name:
            continue
        columns[field_name] = ColumnHeader(
            index=col_index,
            desc=desc,
            type_name=type_name,
            field_name=field_name,
            export_target=export_target,
        )

    data_rows: dict[str, DataRow] = {}
    key_field = next(iter(columns), None)
    for row_number, raw_row in enumerate(rows[HEADER_ROWS:], start=DATA_START_ROW):
        values: dict[str, Any] = {}
        for field_name, header in columns.items():
            values[field_name] = value_at(raw_row, header.index)
        if not any(not is_empty(value) for value in values.values()):
            continue
        row_key = derive_row_key(values, key_field)
        data_rows[row_key] = DataRow(row_number=row_number, key=row_key, values=values)
    return SheetTable(columns=columns, rows=data_rows)


def derive_row_key(values: dict[str, Any], key_field: str | None) -> str:
    if key_field and not is_empty(values.get(key_field)):
        return normalize_cell(values[key_field])
    for value in values.values():
        if not is_empty(value):
            return normalize_cell(value)
    return "<empty>"


def normalize_sheet_key(sheet_name: str) -> str:
    match = SHEET_NAME_PATTERN.search(sheet_name)
    if match:
        return match.group(1).strip().lower()
    return sheet_name.strip().lower()


def normalize_field_name(field_name: Any, desc: Any) -> str:
    if not is_empty(field_name):
        return str(field_name).strip()
    if not is_empty(desc):
        return str(desc).strip()
    return ""


def value_at(row: tuple[Any, ...], index: int) -> Any:
    if index >= len(row):
        return None
    return row[index]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() in EMPTY_VALUES:
        return ""
    return text


def is_empty(value: Any) -> bool:
    return normalize_cell(value) == ""


def format_value(value: Any) -> str:
    text = normalize_cell(value)
    if len(text) > 120:
        return text[:117] + "..."
    return text


def format_report(diffs: list[DiffItem], dir_a: Path, dir_b: Path) -> str:
    lines = [
        "=" * 72,
        "目录 Merge 检查结果",
        f"目录 A（目标）: {dir_a}",
        f"目录 B（来源）: {dir_b}",
        "=" * 72,
    ]
    if not diffs:
        lines.append("未发现目录 B 中存在、但目录 A 未 merge 的内容。")
        lines.append("=" * 72)
        return "\n".join(lines)

    grouped: dict[str, list[DiffItem]] = {}
    for item in diffs:
        grouped.setdefault(item.table, []).append(item)

    for table_name, items in grouped.items():
        lines.append(f"\n[表] {table_name}")
        for item in items:
            lines.append(f"  - 类型: {item.diff_type}")
            lines.append(f"    Sheet: {item.sheet}")
            lines.append(f"    表头/主键: {item.header}")
            lines.append(f"    详情: {item.detail}")
            if item.svn_revision:
                lines.append(
                    "    未 merge 的 SVN 提交: "
                    f"r{item.svn_revision} | {item.svn_author or '-'} | {item.svn_message or '-'}"
                )
            else:
                lines.append("    未 merge 的 SVN 提交: 未能自动定位（请手动 svn log 该文件）")
    lines.append("\n" + "=" * 72)
    lines.append(f"共发现 {len(diffs)} 处未 merge 内容。")
    lines.append("=" * 72)
    return "\n".join(lines)


def attach_svn_info(item: DiffItem, file_path: Path, limit: int) -> None:
    entry = latest_svn_log_entry(file_path, limit=limit)
    if entry:
        item.svn_revision = entry["revision"]
        item.svn_author = entry["author"]
        item.svn_message = entry["message"]


def attach_column_revision(item: DiffItem, file_path: Path, sheet_name: str, field_name: str, limit: int) -> None:
    entry = find_column_introducing_revision(file_path, sheet_name, field_name, limit)
    if entry:
        item.svn_revision = entry["revision"]
        item.svn_author = entry["author"]
        item.svn_message = entry["message"]


def attach_row_revision(item: DiffItem, file_path: Path, sheet_name: str, row_key: str, limit: int) -> None:
    entry = find_row_introducing_revision(file_path, sheet_name, row_key, limit)
    if entry:
        item.svn_revision = entry["revision"]
        item.svn_author = entry["author"]
        item.svn_message = entry["message"]


def attach_cell_revision(
    item: DiffItem,
    file_path: Path,
    sheet_name: str,
    row_key: str,
    field_name: str,
    expected_value: Any,
    limit: int,
) -> None:
    entry = find_cell_introducing_revision(file_path, sheet_name, row_key, field_name, expected_value, limit)
    if entry:
        item.svn_revision = entry["revision"]
        item.svn_author = entry["author"]
        item.svn_message = entry["message"]


def latest_svn_log_entry(file_path: Path, limit: int) -> dict[str, str] | None:
    entries = svn_log_entries(file_path, limit=limit)
    return entries[0] if entries else None


def find_column_introducing_revision(
    file_path: Path,
    sheet_name: str,
    field_name: str,
    limit: int,
) -> dict[str, str] | None:
    entries = svn_log_entries(file_path, limit=limit)
    previous_has_column = None
    for entry in entries:
        table = load_sheet_table_from_revision(file_path, entry["revision"], sheet_name)
        if table is None:
            continue
        has_column = field_name in table.columns
        if previous_has_column is False and has_column:
            return entry
        previous_has_column = has_column
    return entries[0] if entries else None


def find_row_introducing_revision(
    file_path: Path,
    sheet_name: str,
    row_key: str,
    limit: int,
) -> dict[str, str] | None:
    entries = svn_log_entries(file_path, limit=limit)
    previous_has_row = None
    for entry in entries:
        table = load_sheet_table_from_revision(file_path, entry["revision"], sheet_name)
        if table is None:
            continue
        has_row = row_key in table.rows
        if previous_has_row is False and has_row:
            return entry
        previous_has_row = has_row
    return entries[0] if entries else None


def find_cell_introducing_revision(
    file_path: Path,
    sheet_name: str,
    row_key: str,
    field_name: str,
    expected_value: Any,
    limit: int,
) -> dict[str, str] | None:
    entries = svn_log_entries(file_path, limit=limit)
    target = normalize_cell(expected_value)
    previous_value = None
    for entry in entries:
        table = load_sheet_table_from_revision(file_path, entry["revision"], sheet_name)
        if table is None or row_key not in table.rows or field_name not in table.columns:
            current_value = None
        else:
            current_value = normalize_cell(table.rows[row_key].values.get(field_name))
        if previous_value is not None and previous_value != target and current_value == target:
            return entry
        previous_value = current_value
    return entries[0] if entries else None


def load_sheet_table_from_revision(file_path: Path, revision: str, sheet_name: str) -> SheetTable | None:
    content = svn_cat(file_path, revision)
    if content is None:
        return None
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_path = Path(temp_file.name)
        temp_file.write(content)
    try:
        workbook = load_workbook(temp_path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                normalized = normalize_sheet_key(sheet_name)
                actual_name = next(
                    (name for name in workbook.sheetnames if normalize_sheet_key(name) == normalized),
                    None,
                )
                if actual_name is None:
                    return None
                sheet_name = actual_name
            return load_sheet_table(workbook[sheet_name])
        finally:
            workbook.close()
    finally:
        temp_path.unlink(missing_ok=True)


def svn_log_entries(file_path: Path, limit: int) -> list[dict[str, str]]:
    if not is_svn_path(file_path):
        return []
    result = run_command(
        [
            "svn",
            "log",
            f"-l{limit}",
            str(file_path),
        ]
    )
    if result.returncode != 0:
        return []
    return parse_svn_log(result.stdout)


def svn_cat(file_path: Path, revision: str) -> bytes | None:
    if not is_svn_path(file_path):
        return None
    result = run_command(["svn", "cat", "-r", revision, str(file_path)], binary=True)
    if result.returncode != 0:
        return None
    return result.stdout


def is_svn_path(file_path: Path) -> bool:
    result = run_command(["svn", "info", str(file_path)])
    return result.returncode == 0


def parse_svn_log(text: str) -> list[dict[str, str]]:
    entries: list[dict[str, str]] = []
    blocks = re.split(r"\n(?=^-{72}\nr)", text.strip())
    for block in blocks:
        if not block.strip():
            continue
        lines = block.splitlines()
        if not lines:
            continue
        header = lines[0].lstrip("-").strip()
        match = re.match(r"r(\d+)\s+\|\s+([^|]+?)\s+\|\s+([^|]+?)\s+\|\s+(\d+)", header)
        if not match:
            continue
        message_lines = []
        for line in lines[1:]:
            if line.startswith("------"):
                break
            message_lines.append(line)
        entries.append(
            {
                "revision": match.group(1),
                "author": match.group(2).strip(),
                "date": match.group(3).strip(),
                "message": "\n".join(message_lines).strip(),
            }
        )
    return entries


@dataclass(slots=True)
class CommandResult:
    returncode: int
    stdout: bytes | str


def run_command(args: list[str], binary: bool = False) -> CommandResult:
    completed = subprocess.run(args, capture_output=True, check=False)
    stdout = completed.stdout if binary else completed.stdout.decode("utf-8", errors="replace")
    return CommandResult(returncode=completed.returncode, stdout=stdout)


if __name__ == "__main__":
    raise SystemExit(main())
