#!/usr/bin/env python3
import argparse
import json
import os
import tempfile
import zipfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


def cell_content(cell):
    return {
        "coordinate": cell.coordinate,
        "value": cell.value,
        "data_type": cell.data_type,
        "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
        "comment": cell.comment.text if cell.comment else None,
    }


def content_snapshot(sheet):
    return [
        cell_content(cell)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None or cell.hyperlink is not None or cell.comment is not None
    ]


def style_signature(cell):
    side = lambda value: (value.style, value.color.type if value.color else None, value.color.rgb if value.color and value.color.type == "rgb" else None)
    return {
        "font": (cell.font.name, cell.font.sz, cell.font.bold, cell.font.italic, cell.font.color.type if cell.font.color else None, cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None),
        "fill": (cell.fill.fill_type, cell.fill.fgColor.type, cell.fill.fgColor.rgb),
        "border": (side(cell.border.left), side(cell.border.right), side(cell.border.top), side(cell.border.bottom)),
        "alignment": (cell.alignment.horizontal, cell.alignment.vertical, cell.alignment.wrap_text, cell.alignment.text_rotation, cell.alignment.shrink_to_fit),
        "number_format": cell.number_format,
        "protection": (cell.protection.locked, cell.protection.hidden),
    }


def copy_format(source, target):
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def last_nonempty_row(sheet, first_column, last_column):
    for row in range(sheet.max_row, 0, -1):
        if any(sheet.cell(row, column).value is not None for column in range(first_column, last_column + 1)):
            return row
    return 1


def audit(path, sheet_name):
    with zipfile.ZipFile(path) as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise RuntimeError(f"ZIP成员损坏: {bad_member}")
    book = load_workbook(path, data_only=False)
    sheet = book[sheet_name] if sheet_name else book.active
    report = {
        "path": str(path),
        "sheet": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "auto_filter": sheet.auto_filter.ref,
        "merged_ranges": [str(value) for value in sheet.merged_cells.ranges],
        "column_widths": {key: value.width for key, value in sheet.column_dimensions.items()},
        "header_style": {cell.coordinate: style_signature(cell) for cell in sheet[1]},
        "body_style": {cell.coordinate: style_signature(cell) for cell in sheet[2]} if sheet.max_row >= 2 else {},
    }
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))


def format_like(target_path, template_path, output_path, sheet_name, columns):
    target_book = load_workbook(target_path, data_only=False)
    template_book = load_workbook(template_path, data_only=False)
    target = target_book[sheet_name] if sheet_name else target_book.active
    template = template_book[sheet_name] if sheet_name else template_book.active
    before = content_snapshot(target)

    first_column = target[columns.split(":")[0] + "1"].column
    last_column = target[columns.split(":")[-1] + "1"].column
    for column in range(first_column, last_column + 1):
        letter = target.cell(1, column).column_letter
        target.column_dimensions[letter].width = template.column_dimensions[letter].width
        copy_format(template.cell(1, column), target.cell(1, column))

    effective_last_row = last_nonempty_row(target, first_column, last_column)
    for row in range(2, effective_last_row + 1):
        target.row_dimensions[row].height = template.row_dimensions[2].height
        for column in range(first_column, last_column + 1):
            copy_format(template.cell(2, column), target.cell(row, column))
    target.row_dimensions[1].height = template.row_dimensions[1].height
    target.freeze_panes = template.freeze_panes
    target.sheet_view.showGridLines = template.sheet_view.showGridLines
    target.auto_filter.ref = f"{target.cell(1, first_column).coordinate}:{target.cell(effective_last_row, last_column).coordinate}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=output_path.parent, delete=False) as handle:
        temporary_path = Path(handle.name)
    try:
        target_book.save(temporary_path)
        check_book = load_workbook(temporary_path, data_only=False)
        check = check_book[target.title]
        if content_snapshot(check) != before:
            raise RuntimeError("格式修改导致单元格内容、公式、超链接或批注变化")
        for column in range(first_column, last_column + 1):
            if style_signature(check.cell(1, column)) != style_signature(template.cell(1, column)):
                raise RuntimeError(f"表头格式不一致: {check.cell(1, column).coordinate}")
        for row in range(2, effective_last_row + 1):
            for column in range(first_column, last_column + 1):
                if style_signature(check.cell(row, column)) != style_signature(template.cell(2, column)):
                    raise RuntimeError(f"正文格式不一致: {check.cell(row, column).coordinate}")
        if check.freeze_panes != template.freeze_panes:
            raise RuntimeError(f"冻结窗格不一致: {check.freeze_panes} != {template.freeze_panes}")
        with zipfile.ZipFile(temporary_path) as archive:
            if archive.testzip() is not None:
                raise RuntimeError("输出xlsx的ZIP结构损坏")
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    print(f"formatted: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="审计xlsx，或按模板安全复制表格格式")
    subparsers = parser.add_subparsers(dest="command", required=True)
    audit_parser = subparsers.add_parser("audit")
    audit_parser.add_argument("path", type=Path)
    audit_parser.add_argument("--sheet")
    format_parser = subparsers.add_parser("format-like")
    format_parser.add_argument("target", type=Path)
    format_parser.add_argument("template", type=Path)
    format_parser.add_argument("output", type=Path)
    format_parser.add_argument("--sheet")
    format_parser.add_argument("--columns", default="A:G")
    args = parser.parse_args()
    if args.command == "audit":
        audit(args.path, args.sheet)
    else:
        format_like(args.target, args.template, args.output, args.sheet, args.columns)


if __name__ == "__main__":
    main()
